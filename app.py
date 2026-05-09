# -*- coding: utf-8 -*-
import streamlit as st
import datetime
import pandas as pd
from io import BytesIO
import calendar
import sqlite3
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# =========================
# 資料庫功能 (Backend Stats)
# =========================
DB_FILE = 'stats.db'

def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS downloads 
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, 
                  timestamp DATETIME DEFAULT CURRENT_TIMESTAMP, 
                  filename TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS visits 
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, 
                  timestamp DATETIME DEFAULT CURRENT_TIMESTAMP)''')
    conn.commit()
    conn.close()

def log_download(filename):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("INSERT INTO downloads (filename) VALUES (?)", (filename,))
    conn.commit()
    conn.close()

def log_visit():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("INSERT INTO visits (timestamp) VALUES (CURRENT_TIMESTAMP)")
    conn.commit()
    conn.close()

def get_download_stats():
    conn = sqlite3.connect(DB_FILE)
    df = pd.read_sql_query("SELECT timestamp, filename FROM downloads ORDER BY timestamp DESC", conn)
    conn.close()
    return df

def get_visit_stats():
    conn = sqlite3.connect(DB_FILE)
    df = pd.read_sql_query("SELECT timestamp FROM visits ORDER BY timestamp DESC", conn)
    conn.close()
    return df

init_db()

# =========================
# Streamlit 設定與瀏覽計數
# =========================
st.set_page_config(page_title="樂覺製所生命靈數 | Numerology", layout="centered")

if 'has_visited' not in st.session_state:
    log_visit()
    st.session_state['has_visited'] = True

# =========================
# 公用數字處理
# =========================
def reduce_to_digit(n):
    while n > 9:
        n = sum(int(x) for x in str(n))
    return n

def sum_once(n):
    return sum(int(x) for x in str(n))

def format_layers(total):
    mid = sum_once(total)
    return f"{total}/{mid}/{reduce_to_digit(mid)}" if mid > 9 else f"{total}/{mid}"

# =========================
# 生命靈數主命數計算
# =========================
def calculate_life_path_number(birthday):
    date_str = birthday.strftime("%Y%m%d")
    total_sum = sum(int(char) for char in date_str)
    final_num = reduce_to_digit(total_sum)

    if total_sum != final_num and total_sum > 9:
        second_step = sum_once(total_sum)
        if second_step > 9 and second_step != final_num:
            process_str = f"{total_sum} → {second_step} → {final_num}"
        else:
            process_str = f"{total_sum} → {final_num}"
    else:
        process_str = f"{total_sum} → {final_num}"

    return final_num, total_sum, process_str

# =========================
# 流年計算（修正版）
# ✅ 修正重點：
#   1. 用字串逐位相加，不用數字直接加總
#   2. 流年週期以生日當天為起點：
#      查詢日 >= 今年生日 → 流年年份 = 今年
#      查詢日 <  今年生日 → 流年年份 = 去年
#
# 範例：生日 1997/4/25，查詢日 2026/4/1
#   → 2026/4/1 < 2026/4/25 → 流年年份 = 2025
#   → "20250425" → 2+0+2+5+0+4+2+5 = 20 → 2+0 = 2 ✅
# =========================
def get_current_flow_year_base(birthday, query_date):
    cutoff = datetime.date(query_date.year, birthday.month, birthday.day)
    return query_date.year if query_date >= cutoff else query_date.year - 1

def life_year_number_for_date(birthday, query_date):
    base_year = get_current_flow_year_base(birthday, query_date)
    total_str = f"{base_year}{birthday.month:02}{birthday.day:02}"
    return reduce_to_digit(sum(int(x) for x in total_str))

def life_year_numbers_current_and_next(birthday, query_date):
    base_year = get_current_flow_year_base(birthday, query_date)
    next_year = base_year + 1
    current_num = reduce_to_digit(sum(int(x) for x in f"{base_year}{birthday.month:02}{birthday.day:02}"))
    next_num    = reduce_to_digit(sum(int(x) for x in f"{next_year}{birthday.month:02}{birthday.day:02}"))
    return current_num, next_num, base_year

# =========================
# 流年解說
# =========================
def get_year_advice(n):
    advice = {
        1: ("自主與突破之年 (Year of Autonomy & Breakthrough)",
            "容易衝動、單打獨鬥 (Impulsive, fighting alone)",
            "設定清晰目標；在決策前先蒐集意見、給自己緩衝時間。",
            "⭐⭐⭐⭐"),
        2: ("協作與關係之年 (Year of Collaboration & Relationships)",
            "過度迎合、忽略自我 (Over-accommodating, ignoring self)",
            "練習明確表達需求、建立健康邊界；耐心溝通。",
            "⭐⭐⭐"),
        3: ("創意與表達之年 (Year of Creativity & Expression)",
            "分心、情緒起伏 (Distracted, emotional fluctuations)",
            "為創作與學習預留固定時段；公開練習表達。",
            "⭐⭐⭐⭐"),
        4: ("穩定與基礎之年 (Year of Stability & Foundation)",
            "壓力感、僵化完美主義 (Stress, rigid perfectionism)",
            "用『可持續的小步驟』築基礎；為計畫預留彈性。",
            "⭐⭐⭐"),
        5: ("變動與自由之年 (Year of Change & Freedom)",
            "焦躁、衝動決策 (Restless, impulsive decisions)",
            "先設安全網再突破；用短衝 (sprint) 測試新方向。",
            "⭐⭐⭐⭐"),
        6: ("關懷與責任之年 (Year of Care & Responsibility)",
            "過度承擔、忽略自我 (Over-burdened, self-neglect)",
            "把『照顧自己』寫進行程；清楚承諾與界線。",
            "⭐⭐⭐"),
        7: ("內省與學習之年 (Year of Introspection & Learning)",
            "孤立、鑽牛角尖 (Isolation, overthinking)",
            "安排獨處＋定期對談；用寫作/冥想整理解讀。",
            "⭐⭐⭐"),
        8: ("事業與財務之年 (Year of Career & Finance)",
            "過度追求成就、忽略健康情感 (Over-achieving, ignoring health/emotions)",
            "設定績效與復原節奏並行；學會授權與談判。",
            "⭐⭐⭐⭐"),
        9: ("收尾與釋放之年 (Year of Completion & Release)",
            "抗拒結束、情緒回顧 (Resisting endings, emotional nostalgia)",
            "用感恩做結案；做斷捨離，替新循環清出空間。",
            "⭐⭐⭐"),
    }
    return advice.get(n, ("年度主題 (Theme)", "—", "—", "⭐⭐⭐"))

# =========================
# 幸運物件資料
# =========================
lucky_map = {
    1: {"色": "🔴 紅色 (Red)", "水晶": "紅瑪瑙、石榴石", "小物": "原子筆"},
    2: {"色": "🟠 橙色 (Orange)", "水晶": "太陽石、橙月光", "小物": "月亮吊飾"},
    3: {"色": "🟡 黃色 (Yellow)", "水晶": "黃水晶、黃虎眼", "小物": "紙膠帶"},
    4: {"色": "🟢 綠色 (Green)", "水晶": "綠幽靈、孔雀石", "小物": "方形石頭"},
    5: {"色": "🔵 藍色 (Blue)", "水晶": "海藍寶、藍紋瑪瑙", "小物": "交通票卡"},
    6: {"色": "🔷 靛色 (Indigo)", "水晶": "青金石、蘇打石", "小物": "愛心吊飾"},
    7: {"色": "🟣 紫色 (Purple)", "水晶": "紫水晶", "小物": "書籤"},
    8: {"色": "💗 粉色 (Pink)", "水晶": "粉晶、草莓晶", "小物": "鋼筆"},
    9: {"色": "⚪ 白色 (White)", "水晶": "白水晶、白月光", "小物": "小香包"},
    0: {"色": "⚫️ 黑色 (Black)", "水晶": "黑曜石", "小物": "護身符"},
}

# =========================
# 流日指引 & 星等
# =========================
flowing_day_guidance_map = {
    "11/2": "與自己的內在靈性連結，打開心眼從心去看清楚背後的真相。",
    "12/3": "創意的想法和能量正在湧現，用純粹且動聽的方式傳遞出來。",
    "13/4": "讓想法不再只是想像，是時候設法落實到自己的現實生活中。",
    "14/5": "轉化現有的狀態，從固有和凝滯的工作、關係中解脫。",
    "15/6": "會特別渴望與某人深入交談、分享心事。",
    "16/7": "整理內在與學習的好時機，感到精神渙散時，需要讓自己靜下來。",
    "17/8": "會特別想處理與金錢、服務或管理相關的問題。",
    "18/9": "在新階段來臨之前，先學會放下、告別與結束。",
    "19/10/1": "會發現自己比平時更容易接收到來自內在或外在的靈感。",
    "20/2": "內在外在都將迎來翻轉式的改變，洞見更加清晰的真相。",
    "21/3": "今天點子和想法會比平常要多，好好運用溝通和表達來創造。",
    "22/4": "多任務、多變動的一天。保持耐心與行動力。",
    "23/5": "是時候接收新的刺激和變動，考驗自己是否有足夠勇氣。",
    "24/6": "關心自己身邊親近的家人朋友，承諾與責任是今天的主題。",
    "25/7": "專注在自己的事情上，在這當中找回內在的平靜與和諧感。",
    "26/8": "強化自信與擔當，適合接下責任、處理財務、設定下一步策略。",
    "27/9": "透過真理看見真相，有意識地放下是今天的重點。",
    "28/10/1": "有強大顯化力與執行力的日子。保持務實、負責的態度。",
    "29/11/2": "透過傾聽和觀察，從更高智慧層次解讀事情。",
    "30/3": "今天的主題是溝通與協調，運用創意來做包裝和行銷。",
    "31/4": "創造中蘊含結構，靈感需要被規劃來落地。",
    "32/5": "保持靈活和彈性，敞開心釋放和接收愛，有機會突破。",
    "33/6": "用創意、好玩的方式去服務和關愛，釋放壓抑。",
    "34/7": "今日會想獨處反思，注意情緒管控。",
    "35/8": "推進與擴張的日子，結合創意與商業頭腦。",
    "36/9": "在理想與現實之間取得平衡點，透過服務與奉獻幫助他人。",
    "37/10/1": "適時站出來為自己發聲，勇敢展現和展開新的行動。",
    "38/11/2": "運用累積的經驗協助夥伴家人，用風趣方式點出問題。",
    "39/12/3": "聲音和語言具有大能量，用話語去讚美自己和他人。",
    "40/4": "以穩固為前提，更新現有的框架，建立新結構。",
    "41/5": "穩定中尋求自由。突破常規，在變動中保持平衡。",
    "42/6": "規矩紀律需與人際關係並重，考量感性層面。",
    "43/7": "有強大的組織和分析能力，留意情緒控管與說話方式。",
    "44/8": "具強大執行力與影響力，避免固執而忽略他人聲音。",
    "45/9": "運用理性邏輯深入省思，成就自身智慧。",
    "46/10/1": "成為帶動者，展現組織合作能力，聚焦目標。",
    "47/11/2": "扮演穩定可靠的關鍵角色，在重要時刻協助他人。",
    "48/12/3": "在審慎評估下，做出富有創意的決策。",
    "49/13/4": "在穩定基礎下做出取捨，提升到更高境界。",
    "50/5": "變動中隱藏機會，享受這美好的時刻。",
    "51/6": "勇敢面對恐懼和創傷，與自己和解。",
    "52/7": "從核心切入剖析，看見真相。適合獨處深思。",
    "53/8": "有機會創造財富或經驗，保持開放。",
    "54/9": "從漫無目的收斂聚焦，放下並感謝過往。",
    "55/10/1": "極度外放和自我展現，留意是否冒犯。保持專注。",
    "56/11/2": "跳脫二元對立的思維模式，平衡自由與承諾。",
    "57/12/3": "留意內在直覺，答案都在那裡。",
    "58/13/4": "在變動中整合出新流程和規則。",
    "59/14/5": "富有挑戰性的一天，過去所學將迎來轉化。"
}

def get_flowing_day_guidance(flowing_day_str):
    return flowing_day_guidance_map.get(flowing_day_str, "")

def get_flowing_day_star(flowing_day_str):
    star_map = {
        "11/2": "🌟🌟", "12/3": "🌟🌟🌟🌟", "13/4": "🌟🌟🌟🌟", "14/5": "🌟🌟",
        "15/6": "🌟🌟🌟🌟", "16/7": "🌟🌟🌟", "17/8": "🌟🌟🌟🌟🌟", "18/9": "🌟🌟",
        "19/10/1": "🌟🌟🌟🌟", "20/2": "🌟🌟🌟", "21/3": "🌟🌟🌟🌟", "22/4": "🌟🌟🌟",
        "23/5": "🌟🌟🌟🌟", "24/6": "🌟🌟🌟", "25/7": "🌟🌟", "26/8": "🌟🌟🌟🌟🌟",
        "27/9": "🌟🌟🌟", "28/10/1": "🌟🌟🌟🌟🌟", "29/11/2": "🌟🌟🌟", "30/3": "🌟🌟🌟🌟",
        "31/4": "🌟🌟🌟🌟", "32/5": "🌟🌟🌟🌟", "33/6": "🌟🌟🌟", "34/7": "🌟🌟",
        "35/8": "🌟🌟🌟🌟🌟", "36/9": "🌟🌟🌟🌟", "37/10/1": "🌟🌟🌟🌟🌟", "38/11/2": "🌟🌟🌟",
        "39/12/3": "🌟🌟🌟🌟", "40/4": "🌟🌟🌟", "41/5": "🌟🌟🌟🌟", "42/6": "🌟🌟🌟",
        "43/7": "🌟🌟🌟", "44/8": "🌟🌟🌟🌟", "45/9": "🌟🌟🌟", "46/10/1": "🌟🌟🌟🌟",
        "47/11/2": "🌟🌟🌟", "48/12/3": "🌟🌟🌟🌟", "49/13/4": "🌟🌟🌟", "50/5": "🌟🌟🌟🌟",
        "51/6": "🌟🌟", "52/7": "🌟🌟🌟", "53/8": "🌟🌟🌟🌟", "54/9": "🌟🌟",
        "55/10/1": "🌟🌟🌟", "56/11/2": "🌟🌟", "57/12/3": "🌟🌟🌟🌟", "58/13/4": "🌟🌟🌟",
        "59/14/5": "🌟🌟🌟🌟🌟"
    }
    return star_map.get(flowing_day_str, "🌟🌟🌟")

def get_flowing_year_ref(query_date, bday):
    query_date = query_date.date() if hasattr(query_date, "date") else query_date
    cutoff = datetime.date(query_date.year, bday.month, bday.day)
    return query_date.year if query_date >= cutoff else query_date.year - 1

def get_flowing_month_ref(query_date, birthday):
    query_date = query_date.date() if hasattr(query_date, "date") else query_date
    if query_date.day < birthday.day:
        return query_date.month - 1 if query_date.month > 1 else 12
    return query_date.month

# =========================
# 匯出 Excel 樣式
# =========================
def style_excel(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="流年月曆")
        workbook = writer.book
        worksheet = workbook["流年月曆"]
        header_font = Font(size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        for idx, column in enumerate(df.columns):
            max_length = max((len(str(cell)) for cell in df[column]), default=15)
            adjusted_width = max(15, min(int(max_length * 1.2), 100))
            worksheet.column_dimensions[chr(65 + idx)].width = adjusted_width
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            worksheet.row_dimensions[row[0].row].height = 35
    return output

# =========================
# Streamlit 介面
# =========================
st.title("🧭 樂覺製所生命靈數")
st.markdown(
    "在數字之中，我們與自己不期而遇。\n"
    "(In numbers, we meet ourselves unexpectedly.)\n\n"
    "**Be true, be you — 讓靈魂，自在呼吸。(Let the soul breathe freely.)**"
)

# -------- 區塊 A：流年速算 --------
st.subheader("🌟 生命靈數 & 流年速算 (Life Path & Yearly Flow)")
col1, col2 = st.columns([1.2, 1.2])
with col1:
    birthday = st.date_input(
        "請輸入生日 (Birthday)",
        value=datetime.date(1990, 1, 1),
        min_value=datetime.date(1900, 1, 1),
        max_value=datetime.date(2100, 12, 31)
    )
with col2:
    # ✅ 預設值 = 今天
    ref_date = st.date_input(
        "查詢日期 (Query Date)",
        value=datetime.date.today(),
        min_value=datetime.date(1900, 1, 1),
        max_value=datetime.date(2100, 12, 31)
    )

if st.button("計算靈數與流年 (Calculate)"):
    life_num, life_sum, life_process = calculate_life_path_number(birthday)
    lucky_life = lucky_map.get(life_num, {})

    st.markdown("---")
    st.subheader(f"🔮 您的生命靈數主命數：【 {life_num} 】號人")
    st.caption(f"Life Path Number: {life_num}")
    st.caption(f"計算公式 (Formula)：將西元生日數字全部加總 ({birthday.strftime('%Y/%m/%d')})")
    st.text(f"計算過程 (Calculation)：{life_sum} → {life_process}")
    if lucky_life:
        st.info(
            f"✨ **幸運色 (Color)**：{lucky_life.get('色')} ｜ "
            f"**水晶 (Crystal)**：{lucky_life.get('水晶')} ｜ "
            f"**小物 (Item)**：{lucky_life.get('小物')}"
        )

    st.markdown("---")

    today_n = life_year_number_for_date(birthday, ref_date)
    current_n, next_n, base_year = life_year_numbers_current_and_next(birthday, ref_date)
    next_year = base_year + 1

    current_start = datetime.date(base_year, birthday.month, birthday.day)
    current_end   = datetime.date(next_year, birthday.month, birthday.day) - datetime.timedelta(days=1)
    next_start    = datetime.date(next_year, birthday.month, birthday.day)
    next_end      = datetime.date(next_year + 1, birthday.month, birthday.day) - datetime.timedelta(days=1)

    st.markdown("### 📊 流年結果 (Yearly Flow Result)")
    st.write(f"**目前流年數（依查詢日期 {ref_date}）：** {today_n}")
    st.caption(
        f"目前流年週期：{current_start.strftime('%Y/%m/%d')} ～ {current_end.strftime('%Y/%m/%d')}"
    )

    title, challenge, action, stars = get_year_advice(today_n)
    lucky_year = lucky_map.get(today_n, {})

    st.markdown("#### 🪄 流年解說 (Guidance for the Year)")
    st.markdown(
        f"**主題 (Theme)**：{title}  \n"
        f"**運勢指數 (Stars)**：{stars}  \n"
        f"**挑戰 (Challenge)**：{challenge}  \n"
        f"**建議行動 (Action)**：{action}  \n\n"
        f"**幸運顏色 (Color)**：{lucky_year.get('色', '')}  \n"
        f"**建議水晶 (Crystal)**：{lucky_year.get('水晶', '')}"
    )

    with st.expander("查看目前與下一個流年週期的完整解讀 (View current & next flow year)"):
        for label_ch, label_en, num, start_d, end_d in [
            ("目前流年", "Current Flow Year", current_n, current_start, current_end),
            ("下一流年", "Next Flow Year",    next_n,    next_start,    next_end),
        ]:
            t, c, a, s = get_year_advice(num)
            lk = lucky_map.get(num, {})
            st.markdown(
                f"**{label_ch} ({label_en}) → 流年數 {num}**  \n"
                f"📅 週期：{start_d.strftime('%Y/%m/%d')} ～ {end_d.strftime('%Y/%m/%d')}  \n"
                f"• 主題 (Theme)：{t}  \n"
                f"• ⭐：{s}  \n"
                f"• 挑戰 (Challenge)：{c}  \n"
                f"• 建議 (Advice)：{a}  \n"
                f"• 幸運色 / 水晶 (Color/Crystal)：{lk.get('色', '')} / {lk.get('水晶', '')}"
            )

# -------- 區塊 B：流年月曆產生器 --------
st.subheader("📅 產生 1 個月份的『流年月曆』建議表 (Generate Monthly Calendar)")
target_month = st.selectbox(
    "請選擇月份 (Select Month)",
    list(range(1, 13)),
    index=datetime.datetime.now().month - 1
)

if st.button("🎉 產生日曆建議表 (Generate Excel)"):
    target_year_for_calendar = ref_date.year
    _, last_day = calendar.monthrange(target_year_for_calendar, target_month)
    days = pd.date_range(
        start=datetime.date(target_year_for_calendar, target_month, 1),
        end=datetime.date(target_year_for_calendar, target_month, last_day)
    )
    data = []
    for d in days:
        fd_total = sum(int(x) for x in f"{birthday.year}{birthday.month:02}{d.day:02}")
        flowing_day = format_layers(fd_total)
        main_number = reduce_to_digit(fd_total)
        lucky = lucky_map.get(main_number, {})
        guidance = get_flowing_day_guidance(flowing_day)

        year_ref = get_flowing_year_ref(d, birthday)
        fy_total = sum(int(x) for x in f"{year_ref}{birthday.month:02}{birthday.day:02}")
        flowing_year = format_layers(fy_total)

        fm_ref = get_flowing_month_ref(d, birthday)
        fm_total = sum(int(x) for x in f"{birthday.year}{fm_ref:02}{birthday.day:02}")
        flowing_month = format_layers(fm_total)

        data.append({
            "日期 (Date)":      d.strftime("%Y-%m-%d"),
            "星期 (Day)":       d.strftime("%A"),
            "流年 (Year Num)":  flowing_year,
            "流月 (Month Num)": flowing_month,
            "流日 (Day Num)":   flowing_day,
            "運勢指數 (Stars)": get_flowing_day_star(flowing_day),
            "指引 (Guidance)":  guidance,
            "幸運色 (Color)":   lucky.get("色", ""),
            "水晶 (Crystal)":   lucky.get("水晶", ""),
            "幸運小物 (Item)":  lucky.get("小物", "")
        })

    df = pd.DataFrame(data)
    st.dataframe(df, use_container_width=True)
    file_name = f"LuckyCalendar_{target_year_for_calendar}_{str(target_month).zfill(2)}.xlsx"
    if not df.empty:
        output = style_excel(df)
        st.markdown("### 樂覺製所生命靈數")
        st.download_button(
            label="📥 點此下載 Excel (Download)",
            data=output.getvalue(),
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            on_click=log_download,
            args=(file_name,)
        )
    else:
        st.warning("⚠️ 無法匯出 Excel：目前資料為空 (No data to export)")

# =========================
# 後台管理區 (側邊欄)
# =========================
st.sidebar.markdown("---")
st.sidebar.subheader("🔒 管理員專區")
admin_password = st.sidebar.text_input("輸入密碼", type="password")

if admin_password == "admin123":
    st.sidebar.success("已登入")
    stats_df = get_download_stats()
    visits_df = get_visit_stats()

    col_a, col_b = st.sidebar.columns(2)
    with col_a:
        st.sidebar.metric("👀 總瀏覽", len(visits_df))
    with col_b:
        st.sidebar.metric("📥 總下載", len(stats_df))

    st.sidebar.write("---")

    if not visits_df.empty:
        with st.sidebar.expander("查看瀏覽紀錄 (Visits)"):
            st.dataframe(visits_df)

    if not stats_df.empty:
        with st.sidebar.expander("查看下載紀錄 (Downloads)"):
            st.dataframe(stats_df)
elif admin_password:
    st.sidebar.error("密碼錯誤")
